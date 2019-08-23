
#create a excel sheet
# import Workbook
from openpyxl import Workbook
from openpyxl import load_workbook
# create Workbook object
#wb=Workbook()
# set file path
filepath = r"C:\Users\drewb\PycharmProjects\Learn/demo2.xlsx"
# save workbook
#wb.save(filepath)

#add data to sheet


# set file path
#filepath = r"C:\Users\drewb\PycharmProjects\Learn/demo2.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# set value for cell A1=1
sheet['A1'] = 1
sheet['B4'] = 76
sheet['B6'] = 8
# set value for cell B2=2
sheet.cell(row=2, column=2).value = 2
# save workbook
wb.save(filepath)